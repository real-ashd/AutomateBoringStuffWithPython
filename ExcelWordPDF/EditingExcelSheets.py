#Editing Excel Sheets
import openpyxl,os
wb = openpyxl.Workbook()		#Creates a new blank Excel File
# print(wb)
# print(wb.sheetnames)
sheet = wb['Sheet']
print(sheet['A1'].value)
sheet['A1'] = 42			#Use this assign value to the cell specified
sheet['A2'] = "Hello"
print(sheet['A1'].value,sheet['A2'].value)

#Saving the Excel Spreadsheet
wb.save('example2.xlsx')
print("File saved to path : "+os.getcwd()+"\\example2.xlsx")

#To add new worksheet to your workbook
sheet2 = wb.create_sheet()
print(wb.sheetnames)
print(sheet2.title)

#To add sheet at specific position
wb.create_sheet(index=0,title="MyNewWorkSheet")

#Change the name of the sheet
sheet2.title = "Sheet2"
sheet.title = "Sheet1"
print(wb.sheetnames)
#Saving the Excel Spreadsheet
wb.save('example3.xlsx')
print("File saved to path : "+os.getcwd()+"\\example3.xlsx")
