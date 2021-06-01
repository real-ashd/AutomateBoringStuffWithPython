#Reading Excel Files
#Install openpyxl using 'pip install openpyxl'
import openpyxl,os
#print(os.getcwd())
workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
print("-------------------------------------------")

#Workbooks has many sheets inside them
# workbook.get_sheet_names()		#Deprecated Function
# workbook.get_sheet_by_name('Sheet1')
print(workbook.sheetnames)
print(workbook['Sheet1'])
sheet1 = workbook['Sheet1']
print("-------------------------------------------")

#Getting cells from the worksheet
cell = sheet1['A1']
print(cell.value)
print(sheet1['B1'].value)

print(sheet1.cell(row=1,column=5))		#row and column values begin at '1' and not '0'

print("-------------------------------------------")
for i in range(1,7):
	print(i,sheet1.cell(row=1,column=i).value)